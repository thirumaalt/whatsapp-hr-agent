from pathlib import Path
from openpyxl import Workbook, load_workbook

OUTPUT_DIR = Path("output")
EXCEL_FILE = OUTPUT_DIR / "jobs.xlsx"

HEADERS = [
    "Company",
    "Job Title",
    "Email",
    "Location",
    "Experience",
    "Apply Link",
    "Mail Sent",
    "Date"
]


def save_jobs(jobs):
    OUTPUT_DIR.mkdir(exist_ok=True)

    # Create Excel if it doesn't exist
    if not EXCEL_FILE.exists():
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Jobs"

        sheet.append(HEADERS)

        workbook.save(EXCEL_FILE)

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook["Jobs"]

    # Read existing emails
    existing_emails = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2]:
            existing_emails.add(row[2].lower())

    # Add only new jobs
    new_jobs = 0

    for job in jobs:

        email = (job.get("email") or "").lower()

        if not email:
            continue

        if email in existing_emails:
            continue

        sheet.append([
            job.get("company"),
            job.get("title"),
            job.get("email"),
            job.get("location"),
            job.get("experience"),
            job.get("apply_link"),
            "No",
            ""
        ])

        existing_emails.add(email)
        new_jobs += 1

    workbook.save(EXCEL_FILE)

    print(f"Added {new_jobs} new jobs.")
