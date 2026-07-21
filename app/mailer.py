import smtplib
import ssl
import time
from pathlib import Path

from openpyxl import load_workbook

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from app.config import (
    EXCEL_FILE,
    MAIL_TEMPLATE,
    RESUME_PATH,
    MAIL_SUBJECT,
    MAIL_DELAY,
    SMTP_SERVER,
    SMTP_PORT,
    SMTP_USERNAME,
    SMTP_PASSWORD,
    FROM_EMAIL,
)


def send_emails():

    # --------------------------------------------------------
    # Validate Files
    # --------------------------------------------------------

    if not EXCEL_FILE.exists():
        print(f"❌ Excel file not found : {EXCEL_FILE}")
        return

    if not MAIL_TEMPLATE.exists():
        print(f"❌ Mail template not found : {MAIL_TEMPLATE}")
        return

    if not RESUME_PATH.exists():
        print(f"❌ Resume not found : {RESUME_PATH}")
        return

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    html_body = MAIL_TEMPLATE.read_text(encoding="utf-8")

    pending = []

    # --------------------------------------------------------
    # Find Pending Emails
    # --------------------------------------------------------

    for row in range(2, sheet.max_row + 1):

        email = sheet.cell(row=row, column=3).value
        status = sheet.cell(row=row, column=7).value

        if not email:
            continue

        if str(status).strip().lower() == "yes":
            continue

        pending.append((row, email))

    print("=" * 70)
    print("EMAIL SENDER")
    print("=" * 70)
    print(f"Pending Emails : {len(pending)}")
    print("=" * 70)

    if len(pending) == 0:
        print("✅ No pending emails.")
        return

    print(f"Preparing to send {len(pending)} email(s)...")

    # --------------------------------------------------------
    # Gmail SMTP Connection
    # --------------------------------------------------------

    try:

        context = ssl.create_default_context()

        smtp = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)

        smtp.ehlo()
        smtp.starttls(context=context)
        smtp.ehlo()

        smtp.login(
            SMTP_USERNAME,
            SMTP_PASSWORD
        )

        print("\n✅ Gmail SMTP Login Successful\n")

    except Exception as e:

        print("\n❌ SMTP Login Failed")
        print(e)
        return

    # --------------------------------------------------------
    # Send Emails
    # --------------------------------------------------------

    success = 0
    failed = 0

    for row, email in pending:

        try:

            print("=" * 70)
            print(f"Recipient : {email}")

            message = MIMEMultipart()

            message["From"] = FROM_EMAIL
            message["To"] = email
            message["Subject"] = MAIL_SUBJECT

            message.attach(
                MIMEText(
                    html_body,
                    "html"
                )
            )

            with open(RESUME_PATH, "rb") as file:

                attachment = MIMEBase(
                    "application",
                    "octet-stream"
                )

                attachment.set_payload(file.read())

            encoders.encode_base64(attachment)

            attachment.add_header(
                "Content-Disposition",
                f'attachment; filename="{Path(RESUME_PATH).name}"'
            )

            message.attach(attachment)

            print("Sending Email...")

            smtp.sendmail(
                FROM_EMAIL,
                [email],
                message.as_string()
            )

            sheet.cell(
                row=row,
                column=7
            ).value = "Yes"

            workbook.save(EXCEL_FILE)

            success += 1

            print("✅ Email Sent Successfully")

            time.sleep(MAIL_DELAY)

        except Exception as e:

            failed += 1

            print("=" * 70)
            print("❌ Failed")
            print(f"Recipient : {email}")
            print(e)
            print("=" * 70)

    smtp.quit()

    workbook.save(EXCEL_FILE)

    print("\n" + "=" * 70)
    print("EMAIL SUMMARY")
    print("=" * 70)
    print(f"Successful : {success}")
    print(f"Failed     : {failed}")
    print("🎉 Completed.")
    print("=" * 70)


if __name__ == "__main__":
    send_emails()
